"""
DHCD TOPA Weekly Reports PDF to Excel Converter
This script downloads all PDF reports from the DHCD website and converts them to Excel.

Requirements:
pip install requests beautifulsoup4 pdfplumber openpyxl

Usage:
python dhcd_pdf_to_excel.py
"""

import requests
from bs4 import BeautifulSoup
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import os
from urllib.parse import urljoin
from datetime import datetime

# Configuration
BASE_URL = "https://dhcd.dc.gov"
SEARCH_URL = "https://dhcd.dc.gov/publications?after%5Bvalue%5D%5Bdate%5D=01%2F01%2F2015&before%5Bvalue%5D%5Bdate%5D=10%2F01%2F2023&keys=Weekly+Report+on+Tenant+Opportunity&type=All&sort_by=field_date_value&sort_order=ASC"
OUTPUT_DIR = "downloaded_pdfs1"
OUTPUT_EXCEL = "CASD_Weekly_Reports_All1.xlsx"

# Create output directory
os.makedirs(OUTPUT_DIR, exist_ok=True)

def fetch_publication_links():
    """Fetch all publication page links from the search results"""
    print("Fetching publication list...")
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    response = requests.get(SEARCH_URL, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Find all publication links
    publication_links = []
    for link in soup.find_all('a', href=True):
        href = link['href']
        if '/publication/weekly-report-tenant-opportunity' in href:
            full_url = urljoin(BASE_URL, href)
            if full_url not in publication_links:
                publication_links.append(full_url)
                print(f"  Found: {link.get_text(strip=True)[:60]}...")
    
    print(f"\nFound {len(publication_links)} publications")
    return publication_links

def extract_pdf_url(publication_url):
    """Extract PDF download link from a publication page"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    response = requests.get(publication_url, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Look for PDF links
    for link in soup.find_all('a', href=True):
        href = link['href']
        if href.endswith('.pdf') or 'files/dc/sites/dhcd' in href:
            pdf_url = urljoin(BASE_URL, href)
            return pdf_url
    
    return None

def download_pdf(pdf_url, filename):
    """Download a PDF file"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    response = requests.get(pdf_url, headers=headers)
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    with open(filepath, 'wb') as f:
        f.write(response.content)
    
    return filepath

def detect_pdf_format(pdf_path):
    """Detect which format the PDF is in (2016 or 2021)"""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            if 'DHCD CASD Mail Log' in text or 'CASD\'s Weekly Report' in text:
                return '2021'
            elif 'TOPA-Related Filings: Weekly Report' in text:
                return '2016'
    except:
        pass
    return 'unknown'

def extract_pdf_data_2021(pdf_path):
    """Extract data from 2021+ format PDF"""
    data_rows = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            current_category = ""
            current_subcategory = ""
            
            for page in pdf.pages:
                text = page.extract_text()
                if not text:
                    continue
                    
                lines = text.split('\n')
                
                for line in lines:
                    line = line.strip()
                    
                    if not line or 'DHCD CASD Mail Log' in line or 'DATE is during' in line:
                        continue
                    
                    # Main category detection
                    if line.startswith('Conversion -'):
                        current_category = 'Conversion'
                        subcategory_match = re.search(r'Conversion - (.+?) - \(', line)
                        if subcategory_match:
                            current_subcategory = subcategory_match.group(1)
                        continue
                    
                    if line.startswith('Sale and Transfer -'):
                        current_category = 'Sale and Transfer'
                        subcategory_match = re.search(r'Sale and Transfer - \(empty\) - (.+?) \(', line)
                        if subcategory_match:
                            current_subcategory = subcategory_match.group(1)
                        continue
                    
                    # Parse data rows with date pattern
                    date_match = re.match(r'^(\d{2}-\d{2}-\d{4})\s+(.+)$', line)
                    if date_match:
                        date = date_match.group(1)
                        remainder = date_match.group(2).strip()
                        
                        # Extract related address (numeric at end), total units, and sales price
                        related_address = ''
                        total_units = ''
                        sales_price = ''
                        address = remainder
                        
                        # Try to extract numeric values at the end
                        parts = remainder.rsplit(None, 3)
                        if len(parts) >= 2:
                            last_part = parts[-1]
                            second_last = parts[-2] if len(parts) > 1 else ''
                            third_last = parts[-3] if len(parts) > 2 else ''
                            
                            # Pattern: address | related_address | units | price
                            if '$' in last_part or ',' in last_part or (last_part.replace(',', '').isdigit() and len(last_part) > 4):
                                sales_price = last_part.replace('$', '').replace(',', '')
                                if second_last.isdigit():
                                    total_units = second_last
                                    if third_last.isdigit():
                                        related_address = third_last
                                        address = ' '.join(parts[:-3])
                                    else:
                                        address = ' '.join(parts[:-2])
                                else:
                                    address = ' '.join(parts[:-1])
                            # Pattern: address | related_address | units
                            elif second_last.isdigit() and last_part.isdigit() and len(last_part) <= 3:
                                total_units = last_part
                                related_address = second_last
                                address = ' '.join(parts[:-2])
                            # Pattern: address | related_address
                            elif last_part.isdigit() and len(last_part) == 4:
                                related_address = last_part
                                address = ' '.join(parts[:-1])
                        
                        data_rows.append([
                            current_category,
                            current_subcategory,
                            date,
                            address,
                            related_address,
                            total_units,
                            sales_price
                        ])
    except Exception as e:
        print(f"    Error extracting data: {e}")
    
    return data_rows

def extract_pdf_data_2016(pdf_path):
    """Extract data from 2016 format PDF"""
    data_rows = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            current_section = ""
            current_category = ""
            
            for page in pdf.pages:
                tables = page.extract_tables()
                
                if not tables:
                    continue
                
                for table in tables:
                    for row in table:
                        if not row or all(not cell for cell in row):
                            continue
                        
                        row_text = ' '.join([str(cell) if cell else '' for cell in row]).strip()
                        
                        if 'TOPA-Related Filings' in row_text or 'DC Department' in row_text:
                            continue
                        
                        # Detect section headers
                        if row_text.startswith('I.'):
                            current_section = 'I. Conversion Data'
                            continue
                        elif row_text.startswith('II.'):
                            current_section = 'II. Sale & Transfer Data'
                            continue
                        
                        # Detect category lines
                        if '# Received:' in row_text and 'Filing Date' in row_text:
                            category_match = re.search(r'Filing Date\s+(.+?)(?:\s+#|$)', row_text)
                            if category_match:
                                current_category = category_match.group(1).strip()
                            else:
                                parts = row_text.split('Filing Date')
                                if len(parts) > 1:
                                    current_category = parts[1].strip()
                            continue
                        
                        # Extract data rows
                        date_pattern = r'(\d{1,2}/\d{1,2}/\d{4})'
                        date_match = re.search(date_pattern, row_text)
                        
                        if date_match:
                            filing_date = date_match.group(1)
                            remaining = row_text[date_match.end():].strip()
                            
                            address = ''
                            description = ''
                            units = ''
                            sale_price = ''
                            
                            desc_patterns = [
                                r'(Offer of Sale w/contract|Offer of Sale w/o contract|Offer of Sale w/ Contract|Offer of Sale w/o Contract)',
                                r'(Right of First Refusal)',
                                r'(Tenant letter of Interest|Assignment of Rights)',
                                r'(DOPA)',
                                r'\(SFD\)|\(2-4\)|\(5\+\)'
                            ]
                            
                            desc_match = None
                            for pattern in desc_patterns:
                                desc_match = re.search(pattern, remaining, re.IGNORECASE)
                                if desc_match:
                                    break
                            
                            if desc_match:
                                address = remaining[:desc_match.start()].strip()
                                description = remaining[desc_match.start():].strip()
                                
                                price_match = re.search(r'(\d+)/\$?([\d,]+)', description)
                                if price_match:
                                    units = price_match.group(1)
                                    sale_price = price_match.group(2)
                                    description = re.sub(r'\d+/\$?[\d,]+', '', description).strip()
                            else:
                                address = remaining
                            
                            data_rows.append([
                                current_section,
                                current_category,
                                filing_date,
                                address,
                                description,
                                units,
                                sale_price
                            ])
    except Exception as e:
        print(f"    Error extracting data: {e}")
    
    return data_rows

def extract_pdf_data(pdf_path):
    """Extract data from PDF, automatically detecting format"""
    format_type = detect_pdf_format(pdf_path)
    
    if format_type == '2021':
        return extract_pdf_data_2021(pdf_path)
    elif format_type == '2016':
        return extract_pdf_data_2016(pdf_path)
    else:
        # Try both formats
        data = extract_pdf_data_2021(pdf_path)
        if not data:
            data = extract_pdf_data_2016(pdf_path)
        return data

def format_sheet(ws):
    """Format Excel sheet with headers"""
    headers = ['Category', 'Subcategory', 'Date', 'Address', 'Related Address', 'Total Units', 'Sales Price']
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.freeze_panes = 'A2'

def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def main():
    print("=" * 70)
    print("DHCD TOPA Weekly Reports PDF to Excel Converter")
    print("=" * 70)
    print()
    
    # Step 1: Fetch publication links
    publication_links = fetch_publication_links()
    
    if not publication_links:
        print("No publications found. Please check the URL or try again later.")
        return
    
    # Step 2: Download PDFs
    pdf_files = []
    print("\nDownloading PDFs...")
    for i, pub_url in enumerate(publication_links, 1):
        try:
            print(f"{i}/{len(publication_links)} Processing: {pub_url}")
            
            # Extract PDF URL from publication page
            pdf_url = extract_pdf_url(pub_url)
            if not pdf_url:
                print(f"  ⚠ No PDF found on this page")
                continue
            
            # Generate filename from URL
            filename = pub_url.split('/')[-1] + '.pdf'
            
            # Download PDF
            print(f"  Downloading PDF...")
            pdf_path = download_pdf(pdf_url, filename)
            pdf_files.append((pdf_path, pub_url.split('/')[-1]))
            print(f"  ✓ Downloaded: {filename}")
            
        except Exception as e:
            print(f"  ✗ Error: {e}")
    
    print(f"\nSuccessfully downloaded {len(pdf_files)} PDFs")
    
    if not pdf_files:
        print("No PDFs downloaded. Exiting.")
        return
    
    # Step 3: Convert to Excel
    print("\nConverting PDFs to Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title='Summary', index=0)
    summary_ws['A1'] = 'CASD Weekly Reports Summary'
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws['A3'] = 'Report'
    summary_ws['B3'] = 'Total Records'
    summary_ws['A3'].font = Font(bold=True)
    summary_ws['B3'].font = Font(bold=True)
    
    # Process each PDF
    summary_row = 4
    for i, (pdf_path, sheet_name) in enumerate(pdf_files, 1):
        try:
            print(f"{i}/{len(pdf_files)} Converting: {sheet_name}")
            
            # Extract data
            data_rows = extract_pdf_data(pdf_path)
            
            if not data_rows:
                print(f"  ⚠ No data extracted")
                continue
            
            # Create sheet (truncate name if too long)
            safe_sheet_name = sheet_name[:31]  # Excel limit
            ws = wb.create_sheet(title=safe_sheet_name)
            format_sheet(ws)
            
            # Add data
            for row in data_rows:
                ws.append(row)
            
            # Format
            auto_adjust_columns(ws)
            
            # Update summary
            summary_ws[f'A{summary_row}'] = safe_sheet_name
            summary_ws[f'B{summary_row}'] = len(data_rows)
            summary_row += 1
            
            print(f"  ✓ Extracted {len(data_rows)} records")
            
        except Exception as e:
            print(f"  ✗ Error: {e}")
    
    # Format summary sheet
    summary_ws.column_dimensions['A'].width = 50
    summary_ws.column_dimensions['B'].width = 15
    
    # Save Excel file
    wb.save(OUTPUT_EXCEL)
    print(f"\n{'=' * 70}")
    print(f"✓ SUCCESS! Excel file created: {OUTPUT_EXCEL}")
    print(f"{'=' * 70}")
    print(f"\nTotal sheets: {len(wb.sheetnames)}")
    print(f"Total records: {sum([ws[f'B{i}'].value or 0 for i in range(4, summary_row)])}")

if __name__ == "__main__":
    main()